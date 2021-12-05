import time
import os
from openpyxl import Workbook
from openpyxl import load_workbook

def verificar(entero):
    if entero != int:
        while entero.isdecimal() == False:
            print("Error, debe ingresar un número")
            entero = input("Ingrese un número: ")
    entero = int(entero)
    return entero

def check(ingreso):
    while ingreso == "" or ingreso.isdecimal() == True:
        print("Error de campo, debe ingresar un nombre válido")
        ingreso = input("Ingrese nuevamente: ")
    ingreso = ingreso.rstrip(",.:;#$%&/(=?¡+´¨{_-{|}[]")
    ingreso = ingreso.capitalize()
    return ingreso

def alcanza(val1,val2):
    while val1 < val2:
        print("No alcanza - ¿Ingresaste bien el pago?")
        val1 = input("El cliente abona con $: ")
        val1 = verificar(val1)
    val1 = int(val1)
    return val1

combo_s = 650
combo_d = 700
combo_t = 800
postre = 250
total_encargado = 0

print("Bienvenido a McDowell's")
encargado = input("Ingrese su nombre encargad@: ")
encargado = check(encargado)

file = open("registro.txt","a")
file.write("IN " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + " Encargad@: " + encargado + "\n")
file.close()

try:
    wb = load_workbook("pruebaTP.xlsx")
    ws = wb.active
    ws.append(["","","","","","",""])
    wb.save("pruebaTP.xlsx")
except FileNotFoundError:    
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Cliente"
    ws['B1'] = "Fecha"
    ws['C1'] = "Combo S"
    ws['D1'] = "Combo D" 
    ws['E1'] = "Combo T" 
    ws['F1'] = "Flurby"
    ws['G1'] = "Total" 


while True:

    print("""
    1 – Ingreso de nuevo pedido
    2 – Cambio de turno
    3 – Apagar sistema 
    """)

    opcion = input("Ingrese la opción deseada: ")

    if opcion == "1":
        cliente = input("¿Cuál es el nombre del cliente?: ")
        cliente = check(cliente)
        
        q_combo_s = input("Ingrese cantidad Combo S: ")
        q_combo_s = verificar(q_combo_s)
        q_combo_d = input("Ingrese cantidad Combo D: ")
        q_combo_d = verificar(q_combo_d)
        q_combo_t = input("Ingrese catnidad Combo T: ")
        q_combo_t = verificar(q_combo_t)
        q_postre = input("Ingrese Cantidad Flurby: ")
        q_postre = verificar(q_postre)
        
        total = q_combo_s * combo_s + q_combo_d * combo_d + q_combo_t * combo_t + q_postre * postre
        print("El total es: ", total)
        pago = input("Abona con $: ")
        pago = verificar(pago)
        pago = alcanza(pago,total)
        vuelto = pago-total
        print("Vuelto $ =",vuelto)
        time.sleep(2)

        total_encargado += total

        ws.append([cliente, time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) , q_combo_s , q_combo_d , q_combo_t , q_postre , total])
        
        wb.save("pruebaTP.xlsx")

    elif opcion == "2":
        file = open("registro.txt","a")
        file.write("OUT " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + " Encargad@: " + encargado+ " $" + str(total_encargado) + "\n")
        file.write("##################################################################"+"\n")
        file.close()

        ws.append(["","","","","","",""])
        wb.save("pruebaTP.xlsx")

        print("Bienvenido a McDowell's")
        encargado = input("Ingrese su nombre encargad@: ")
        encargado = check(encargado)

        total_encargado = 0
        
        file = open("registro.txt","a")
        file.write("IN " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + " Encargad@: " + encargado + "\n")
        file.close()
        print("Bienvenido, "+encargado+"!")
        time.sleep(2)

    elif opcion == "3":
        print("Gracias por usar el sistema!")
        file = open("registro.txt","a")
        file.write("OUT " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + " Encargad@: " + encargado+ " $" + str(total_encargado) + "\n")
        file.write("##################################################################"+"\n")
        file.close()
        
        time.sleep(2)

        os.system("cls") 

        break

    else:
        print("Opción inexistente, intente nuevamente: ")
        time.sleep(2)

    os.system("cls")   
        
